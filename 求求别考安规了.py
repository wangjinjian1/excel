import json, re, os, requests, random, time
from urllib.parse import parse_qs, urlparse
from collections import defaultdict
from openpyxl import load_workbook
from urllib3 import disable_warnings

disable_warnings()

patternTitle = re.compile('[（）。，！,.() /《》<>、：:;；]')

def getKSTM(examId='', token='', filepath='233.txt'):
    qq = defaultdict()
    if examId == '':
        with open(filepath, 'r', encoding='utf-8') as f:
            questions = json.load(f)['data']['question']
    else:
        questions = getTikuById(examId, token)
    for q in questions:
        qq[int(q['SERIAL_NUMBER'])] = patternTitle.sub('', q['QUESTION_CONTENT']).strip().replace('　　', '').replace('　',
                                                                                                                    '')
    return qq


def handleques(questions):
    qq = defaultdict()
    for q in questions:
        qq[int(q['SERIAL_NUMBER'])] = patternTitle.sub('', q['QUESTION_CONTENT']).strip().replace('　　', '').replace('　',
                                                                                                                    '')
    return qq


def charToIndex(s):
    return ord(s) - 62


# type 1 判断  type 2 单选  type 3 多选
def initTiKu(excelpath='tiku'):
    tiku = defaultdict(dict)
    for excel in os.listdir(excelpath):
        if not excel.endswith('xlsx'):
            continue
        print(os.path.join(excelpath, excel))
        wb = load_workbook(os.path.join(excelpath, excel))
        ws = wb.active
        for i in range(2, ws.max_row + 1):

            if ws.cell(row=i, column=1) == None or ws.cell(row=i, column=1).value == '':
                break
            title = patternTitle.sub('', ws.cell(row=i, column=1).value).replace('　　', '').replace('　', '')
            type = ws.cell(row=i, column=2).value.strip()
            answer = ws.cell(row=i, column=11).value.strip()
            tiku[title]['answer'] = answer
            if type == '判断题':
                tiku[title]['type'] = 1
                if ws.cell(row=i, column=charToIndex(answer)).value == None:
                    tiku[title]['content'] = f'{answer}竟然没答案，愚蠢'
                else:
                    tiku[title]['content'] = ws.cell(row=i, column=charToIndex(answer)).value.strip()
            elif type == '单选题':
                tiku[title]['type'] = 2
                if ws.cell(row=i, column=charToIndex(answer)).value == None:
                    tiku[title]['content'] = f'{answer}竟然没答案，愚蠢'
                else:
                    tiku[title]['content'] = ws.cell(row=i, column=charToIndex(answer)).value.strip()
            elif type == '多选题':
                tiku[title]['type'] = 3
                contents = []
                for j in answer:
                    if ws.cell(row=i, column=charToIndex(j)).value == None:
                        contents.append(f'{j}竟然没答案，愚蠢')
                    else:
                        contents.append(ws.cell(row=i, column=charToIndex(j)).value.strip())
                tiku[title]['content'] = ' | '.join(contents)
    with open('tiku.json', 'w+', encoding='utf-8') as f:
        json.dump(tiku, f, ensure_ascii=False)


def getTiku():
    with open('tiku.json', 'r', encoding='utf-8') as f:
        tiku = json.load(f)
    return tiku


# 考试
def printAnswer(examId=''):
    # qq = getKSTM(examId)
    qq = getKSTM('')
    tiku = getTiku()
    for k, v in qq.items():
        print(k, tiku[v]['answer'], tiku[v]['content'])


# 进阶模拟答题
def auto1(examId, token, usetime=random.randint(2000, 2500), check=False):
    qq = getTikuById(examId, token)
    ans = []
    ansre = []
    ansre1 = []
    for index, q in enumerate(qq):
        ans.append({
            "SerialNumber": q['SERIAL_NUMBER'],
            "AnswerValue": q['RIGHT_ANSWERS']
        })
        ansre.append(q['RIGHT_ANSWERS'])
    if check:
        tiku = getTiku()
        for v in qq.values():
            if v in tiku:
                ansre1.append(tiku[v]['answer'])
    data = {'json': json.dumps(ans, separators=(',', ':')),
            'Token': token,
            'TableId': examId,
            'UseTime': usetime,
            'Type': 'Mock'}
    res = requests.post('https://aj.erow.cn:8443/AJGKAPP/API2/EDU_EXERCISE/CommitMockExamAnswer.ashx',
                        params={'t': int(time.time() * 1000)}, data=data,
                        verify=False)
    print(res.text)
    return ansre, ansre1


# 自动答题
# usetime  2400s-2500s
def auto(examId, token, TableId='', usetime=random.randint(2000, 2500)):
    ans = []
    qq = getKSTM(examId)
    tiku = getTiku()
    for k, v in qq.items():
        if v in tiku:
            ans.append({
                "SerialNumber": k,
                "AnswerValue": tiku[v]['answer']
            })
        else:
            ans.append({
                "SerialNumber": k,
                "AnswerValue": "A"
            })

    data = {'json': json.dumps(ans, separators=(',', ':')),
            'Token': token,
            'TableId': TableId if TableId else examId,
            'UseTime': usetime,
            'Type': 'Mock'}
    res = requests.post('https://aj.erow.cn:8443/AJGKAPP/API2/EDU_EXERCISE/CommitMockExamAnswer.ashx',
                        params={'t': int(time.time() * 1000)}, data=data,
                        verify=False)
    print(res.text)


def getJson():
    d = {}
    with open('strtojson.txt', 'r', encoding='utf-8') as f:
        for line in f.readlines():
            arr = line.split(':')
            d[arr[0].strip()] = arr[1].strip()
    print(d)
    return d


def getTikuById(examId, token):
    query = {'t': int(time.time() * 1000),
             'Token': token,
             'questionQty': 195,
             'SpecialType': 353,
             # 'SpecialType': 344,
             'libraryId': 523,
             'libraryType': '',
             'Type': 'Mock',
             'examId': examId}
    res = requests.get('https://aj.erow.cn:8443/AJGKAPP/API2/EDU_EXERCISE/GetMockQuestionList2.ashx', params=query,
                       verify=False)
    print(res.json())
    return res.json()['data']['question']


def kaoshi(excelpath='tiku1'):
    initTiKu(excelpath)
    printAnswer('')


def automoni(a):
    parastr = urlparse(a).query
    query = parse_qs(parastr)
    token = query['Token'][0]
    examId = query['examId'][0]
    print(token, examId)
    a1,a2=auto1(examId=examId, token=token,check=True)
    if a1==a2:
        print('1')
    else:
        print('2')


if __name__ == '__main__':
    # 考试
    # kaoshi('tiku')
    # 模拟
    url = ''
    automoni(url)
