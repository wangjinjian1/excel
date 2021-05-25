from openpyxl import workbook, load_workbook
import random

dic = {
    1: 'A',
    2: 'B',
    3: 'C',
    4: 'D',
    12: 'AB',
    13: 'AC',
    14: 'AD',
    23: 'BC',
    24: 'BD',
    34: 'CD',
    123: 'ABC',
    124: 'ABD',
    134: 'ACD',
    234: 'BCD',
    1234: 'ABCD',
}

dict2 = {
    '正确': 'A',
    '错误': 'B',
    'A': 'A',
    'B': 'B'
}

list = ['题干', '题型', '选项A', '选项B', '选项C', '选项D', '选项E', '选项F', '选项G', '选项H', '正确答案', '解析', '章节', '难度']
list1 = ['晚安', '有手就行', '你也太强了吧', '我是邱曙光，我最帅']


def handeleExcel(ws, kwargs):
    ws.cell(row=1, column=1).value = '标题'
    ws.merge_cells(start_row=1, end_row=1, start_column=2, end_column=8)
    if 'title' in kwargs.keys():
        ws.cell(row=1, column=2).value = kwargs['title']
    ws.cell(row=2, column=1).value = '描述'
    ws.merge_cells(start_row=2, end_row=2, start_column=2, end_column=8)
    if 'description' in kwargs.keys():
        ws.cell(row=2, column=2).value = kwargs['description']
    ws.cell(row=3, column=1).value = '用时'
    if 'time' in kwargs.keys():
        ws.cell(row=3, column=2).value = kwargs['time']
    for i in range(1, 9):
        ws.cell(row=4, column=i).value = list[i - 1]


def fun1(path, **kwargs):
    wb = load_workbook(path)
    ws = wb.active
    for i in range(5, ws.max_row + 1):
        value = ws.cell(row=i, column=8)
        ws.cell(row=i, column=9).value = dic[value.value]
        ws.cell(row=i, column=7).value = list1[random.randint(0, 3)]
    ws.delete_cols(8, 1)
    handeleExcel(ws, kwargs)
    wb.save('@' + path)


def fun2(path, **kwargs):
    wb = load_workbook(path)
    ws = wb.active
    for i in range(5, ws.max_row + 1):
        value = ws.cell(row=i, column=2)
        ws.cell(row=i, column=9).value = dict2[value.value]
        ws.cell(row=i, column=8).value = list1[random.randint(0, 3)]
    ws.delete_cols(2, 1)
    handeleExcel(ws, kwargs)
    wb.save('@' + path)


def fun3(path, **kwargs):
    wb = load_workbook(path)
    ws = wb.active
    for i in range(5, ws.max_row + 1):
        value = ws.cell(row=i, column=2)
        arrs = value.value.split('$;$')
        if len(arrs) > 2:
            for index in range(6, 6 + len(arrs)):
                ws.cell(row=i, column=index).value = arrs[index - 6]
            ws.cell(row=i, column=11).value = ws.cell(row=i, column=3).value
        else:
            ws.cell(row=i, column=11).value = dict2[ws.cell(row=i, column=3).value]
        if len(ws.cell(row=i, column=4).value) > 0:
            ws.cell(row=i, column=10).value = ws.cell(row=i, column=4).value
        else:
            ws.cell(row=i, column=10).value = list1[random.randint(0, 3)]
    ws.delete_cols(2, 3)
    handeleExcel(ws, kwargs)
    wb.save('@' + path)


if __name__ == '__main__':
    fun3('抄表核算收费员网大版题库.xlsx', title='抄表核算收费员网大版题库', description='2333', time=60)
