from openpyxl import Workbook
import re
import os

wb = Workbook()
ws = wb.active
f = open(f'gwxt.txt', 'w+', encoding='utf8')
path_dir='/Users/wangjinjian/Downloads/tiku'
topic = re.compile('(?<=topic:").*(?=",topicLevel)')
topicOption = re.compile('(?<=topicOption:").*(?=",topicKey)')
topicKey = re.compile('(?<=topicKey:").*(?=",keywordsContent)')
splitPa = re.compile(r'\$;(?:\\n)?\$')
ws.append(['题目','选项A','选项B','选项C','选项D','选项E','选项F','答案'])
cnt=1
total=0
for file in os.listdir(path_dir):
    file_path=os.path.join(path_dir,file)
    with open(file_path,'r',encoding='utf8') as f1:
        text=f1.read()
    topic_ = topic.findall(text)
    topicOption_ = topicOption.findall(text)
    topicKey_ = topicKey.findall(text)
    total+=len(topic_)
    for i in range(len(topic_)):
        cnt+=1
        ans = [a.strip().replace('\\n', '').replace('\\t', '') for a in splitPa.split(topicOption_[i])]
        f.write(f'{i + 1}  {topic_[i]} \n')
        ws.cell(row=cnt, column=1).value = topic_[i]
        if len(ans) == 1:
            f.write('A.正确  B.错误')
            ws.cell(row=cnt, column=2).value = '正确'
            ws.cell(row=cnt, column=3).value = '错误'
        else:
            for j in range(len(ans)):
                ws.cell(row=cnt, column=j + 2).value = ans[j][1:]
                f.write(f'{chr(j + 65)}. {ans[j][1:]} \t')
        f.write('\n')
        f.write(f'{topicKey_[i]} \n')
        ws.cell(row=cnt, column=8).value = topicKey_[i]
        f.write('\n')
wb.save('gwxt.xlsx')
f.close()
print(total)
