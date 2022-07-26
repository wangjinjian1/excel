import json, re, os, requests, random, time
from collections import defaultdict
from openpyxl import load_workbook

patternTitle = re.compile('[（）。，！,.() /《》<>、：:;；]')
patternsplit = re.compile(r'(?:\$;\$)|(?:\$\$)')
topicPa = re.compile('(?<=topic:")(.*?)(?=")')
topicKeyPa = re.compile('(?<=topicKey:")(.*?)(?=")')
topicOptionPa = re.compile('(?<=topicOption:")(.*?)(?=")')

d = {
    'A': '正确',
    'B': '错误'
}


def intiTiku(excelDic):
    diccc = defaultdict(dict)
    files = os.listdir(excelDic)
    for file in files:
        excelPath = os.path.join(excelDic, file)
        if not file.endswith('.xlsx') or file.startswith('~$'):
            continue
        # 判断是否选择题
        isSele = file.find('选择') >= 0
        wb = load_workbook(excelPath)
        ws = wb.active
        if isSele:
            for i in range(1, ws.max_row + 1):
                c1 = ws.cell(row=i, column=1).value
                if c1 == None or c1 == '':
                    break
                c2 = patternTitle.sub('', ws.cell(row=i, column=2).value).replace(' ', '')
                c3 = ws.cell(row=i, column=3).value
                c4 = ws.cell(row=i, column=4).value.strip().replace(" ", "")
                ss = patternsplit.split(c3)
                ans = []
                for j in c4:
                    ans.append(ss[charToIndex(j)].strip().replace('\n', ''))
                if c2 in diccc:
                    print(c2)
                diccc[c2]['ans'] = ans
                diccc[c2]['sele'] = c4
        else:
            for i in range(1, ws.max_row + 1):
                c1 = ws.cell(row=i, column=1).value
                if c1 == None or c1 == '':
                    break
                c2 = patternTitle.sub('', ws.cell(row=i, column=2).value)
                c4 = ws.cell(row=i, column=4).value.strip().replace(" ", "")
                if c2 in diccc:
                    print(c2)
                diccc[c2]['ans'] = charToJud(c4)
                diccc[c2]['sele'] = c4
    print(len(diccc))
    with open('pp.json','w+',encoding='utf-8') as f:
        json.dump(diccc,f,ensure_ascii=False)


def analysisT():
    with open('233.txt', 'r', encoding='utf-8') as f:
        s = f.read()
    topic = topicPa.finditer(s)
    topicKey = topicKeyPa.finditer(s)
    topicOption = topicOptionPa.finditer(s)
    for i in zip(topic, topicOption, topicKey):
        t1 = i[0].group()
        t2 = i[2].group()
        t3 = i[1].group().split('$;$')
        ans = []
        for i in t2:
            ans.append(t3[charToIndex(i)])
        print(t2, '|'.join(ans), patternTitle.sub('', t1))


def charToIndex(s):
    return ord(s) - 65


def charToJud(s):
    return d[s]

if __name__ == '__main__':
    # analysisT()
    intiTiku('111')
