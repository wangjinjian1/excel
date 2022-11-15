from openpyxl import load_workbook, Workbook
import os

path = '/Users/wangjinjian/Desktop/关于开展公司2022年大数据应用技能竞赛模拟测试的通知/2022大数据应用技能竞赛参考样题..xlsx'
dirname = os.path.dirname(path)
filename = os.path.basename(path)
wbnw = Workbook()
wsnw = wbnw.active
wsnw.append(['题型', '题目', 'A', 'B', 'C', 'D', '答案'])
wb = load_workbook(path)
ws = wb.active
res = []
for i in range(2, ws.max_row + 1):
    r = []
    tm = ws.cell(row=i, column=3).value
    tx = ws.cell(row=i, column=2).value
    xx = ws.cell(row=i, column=4).value
    da = ws.cell(row=i, column=5).value
    if tm == None or tx == None or xx == None or da == None:
        continue
    arrs = [arr.strip('.、 ').replace('$;$','') for arr in xx.split('\n')]
    if len(arrs)<4:
        arrs+=['']*(4-len(arrs))
    if len(da)==1:
        tx='单选题'
    else:
        tx='多选题'
    r.append(tx)
    r.append(tm)
    r.extend(arrs)
    r.append(da.replace('、','').strip())
    res.append(r)

for r in res:
    wsnw.append(r)
wbnw.save(os.path.join(dirname, f'@{filename}'))
