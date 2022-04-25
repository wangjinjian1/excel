from openpyxl import workbook, load_workbook, Workbook
import random, os

dicSelect = {
    1: 'A',
    2: 'B',
    3: 'C',
    4: 'D',
    12: 'AB',
    13: 'AC',
    14: 'AD',
    23: 'BC',
    24: 'BD',
    34: 'CD',
    123: 'ABC',
    124: 'ABD',
    134: 'ACD',
    234: 'BCD',
    1234: 'ABCD',
    25:'BE',
    1345:'ACDE',
    145:'ADE',
    2345:'BCDE',
    135:'ACE'

}

dicJudge = {
    '正确': 'A',
    '错误': 'B',
    'A': 'A',
    'B': 'B'
}

listtitle = ['题干（必填）', '题型（必填）', '选项A', '选项B', '选项C', '选项D', '选项E', '选项F', '选项G', '选项H', '正确答案（必填）', '解析', '章节', '难度']

listType = ['单选题', '多选题', '判断题']


def modifyAnswer(file, column):
    wb = load_workbook(file)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        answer = ws.cell(row=row, column=column).value
        ws.cell(row=row, column=column).value = dicSelect[answer]
    wb.save(file)


def handeleExcel(ws):
    lenlist = len(listtitle)
    for i in range(1, 1 + lenlist):
        ws.cell(row=1, column=i).value = listtitle[i - 1]
        answer = ws.cell(row=i, column=8)


# 安规选择题
def fun1(path):
    wb = load_workbook(path)
    ws = wb.active
    newwb = Workbook()
    newws = newwb.active
    handeleExcel(newws)
    for row in range(2, ws.max_row + 1):
        newws.cell(row=row, column=1).value = ws.cell(row=row, column=5).value
        answer = ws.cell(row=row, column=11).value
        newws.cell(row=row, column=11).value = dicSelect[answer]
        if answer <= 4:
            newws.cell(row=row, column=2).value = listType[0]
        else:
            newws.cell(row=row, column=2).value = listType[1]
        for i in range(3, 7):
            newws.cell(row=row, column=i).value = ws.cell(row=row, column=i + 3).value
    newwb.save('@' + path)


# 安规判断题
def fun2(path):
    wb = load_workbook(path)
    ws = wb.active
    newwb = Workbook()
    newws = newwb.active
    handeleExcel(newws)
    for row in range(2, ws.max_row + 1):
        newws.cell(row=row, column=1).value = ws.cell(row=row, column=4).value
        answer = ws.cell(row=row, column=5).value
        newws.cell(row=row, column=11).value = dicJudge[answer]
        newws.cell(row=row, column=2).value = listType[2]
        newws.cell(row=row, column=3).value = '正确'
        newws.cell(row=row, column=4).value = '错误'
    newwb.save('@' + path)


# 专业
def fun33(path):
    wb = load_workbook(path)
    ws = wb.active
    newwb = Workbook()
    newws = newwb.active
    handeleExcel(newws)
    for row in range(5, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == None:
            break
        # 题干
        newws.cell(row=row, column=1).value = ws.cell(row=row, column=4).value
        # 答案
        newws.cell(row=row, column=11).value = ws.cell(row=row, column=6).value
        # 题型
        newws.cell(row=row, column=2).value = ws.cell(row=row, column=3).value
        # 解析
        newws.cell(row=row, column=12).value = ws.cell(row=row, column=7).value
        values = ws.cell(row=row, column=5).value
        print(values)
        if values == None:
            arrs = []
        else:
            if values.find('$;$') != -1:
                arrs = values.split('$;$')
            elif values.find('$ ; $') != -1:
                arrs = values.split('$ ; $')
        # 选项
        if len(arrs) > 2:
            for index in range(3, len(arrs) + 3):
                newws.cell(row=row, column=index).value = arrs[index - 3]
        else:
            newws.cell(row=row, column=3).value = '正确'
            newws.cell(row=row, column=4).value = '错误'
    newwb.save('@' + path)


# 专业
def fun3(path, ab=True):
    wb = load_workbook(path)
    ws = wb.active
    newwb = Workbook()
    newws = newwb.active
    handeleExcel(newws)
    for row in range(3, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == None:
            break
        # 题干
        newws.cell(row=row, column=1).value = ws.cell(row=row, column=7).value
        # 答案
        newws.cell(row=row, column=11).value = ws.cell(row=row, column=9).value
        # 题型
        newws.cell(row=row, column=2).value = ws.cell(row=row, column=6).value
        # 解析
        newws.cell(row=row, column=12).value = ws.cell(row=row, column=13).value
        values = ws.cell(row=row, column=8).value
        if values == None:
            arrs = []
        else:
            if values.find('$;$') != -1:
                arrs = values.split('$;$')
            elif values.find('$ ; $') != -1:
                arrs = values.split('$ ; $')
        # 选项
        if len(arrs) > 2:
            for index in range(3, len(arrs) + 3):
                newws.cell(row=row, column=index).value = arrs[index - 3]
        else:
            newws.cell(row=row, column=3).value = '正确'
            newws.cell(row=row, column=4).value = '错误'
    if ab:
        paths = path.split('/')
        lenpath = len(paths)
        fpath = '/'
        for i in range(lenpath - 1):
            fpath = os.path.join(fpath, paths[i])
        fpath = os.path.join(fpath, '@' + paths[lenpath - 1])
        newwb.save(fpath)
    else:
        newwb.save('@' + path)


def RBFight(path, ab=True):
    wb = load_workbook(path)
    ws1 = wb['多选题']
    ws2 = wb['单选择']
    ws3 = wb['判断题']
    newwb = Workbook()
    newws = newwb.active
    handeleExcel(newws)
    cnt = 0
    for row in range(2, ws1.max_row + 1):
        if ws1.cell(row=row, column=2).value == None:
            break
        cnt += 1
        print(ws1.cell(row=row, column=2).value)
        newws.cell(row=row, column=1).value = ws1.cell(row=row, column=2).value
        newws.cell(row=row, column=2).value = '多选题'
        newws.cell(row=row, column=3).value = ws1.cell(row=row, column=3).value[2:]
        newws.cell(row=row, column=4).value = ws1.cell(row=row, column=4).value[2:]
        newws.cell(row=row, column=5).value = ws1.cell(row=row, column=5).value[2:]
        newws.cell(row=row, column=6).value = ws1.cell(row=row, column=6).value[2:]
        newws.cell(row=row, column=11).value = ws1.cell(row=row, column=7).value
    cnt2 = cnt
    for row in range(2, ws2.max_row + 1):
        if ws2.cell(row=row, column=2).value == None:
            break
        print(ws2.cell(row=row, column=2).value)
        cnt2 += 1
        newws.cell(row=row + cnt, column=1).value = ws2.cell(row=row, column=2).value
        newws.cell(row=row + cnt, column=2).value = '单选题'
        newws.cell(row=row + cnt, column=3).value = ws2.cell(row=row, column=3).value[2:]
        newws.cell(row=row + cnt, column=4).value = ws2.cell(row=row, column=4).value[2:]
        newws.cell(row=row + cnt, column=5).value = ws2.cell(row=row, column=5).value[2:]
        newws.cell(row=row + cnt, column=6).value = ws2.cell(row=row, column=6).value[2:]
        newws.cell(row=row + cnt, column=11).value = ws2.cell(row=row, column=7).value
    for row in range(2, ws3.max_row + 1):
        if ws3.cell(row=row, column=2).value == None:
            break
        print(ws3.cell(row=row, column=2).value)
        newws.cell(row=row + cnt2 - 1, column=1).value = ws3.cell(row=row, column=2).value
        newws.cell(row=row + cnt2 - 1, column=2).value = '判断题'
        newws.cell(row=row + cnt2 - 1, column=3).value = ws3.cell(row=row, column=3).value[2:]
        newws.cell(row=row + cnt2 - 1, column=4).value = ws3.cell(row=row, column=4).value[2:]
        newws.cell(row=row + cnt2 - 1, column=11).value = ws3.cell(row=row, column=5).value
    if ab:
        paths = path.split('/')
        lenpath = len(paths)
        fpath = '/'
        for i in range(lenpath - 1):
            fpath = os.path.join(fpath, paths[i])
        fpath = os.path.join(fpath, '@' + paths[lenpath - 1])
        newwb.save(fpath)
    else:
        newwb.save('@' + path)


if __name__ == '__main__':
    # fun2('配电判断题.xlsx')
    fun1('配电选择题.xlsx')
    # fun3('抄表核算收费员网大版题库.xlsx')
    # fun3('/Users/wangjinjian/Downloads/用电监察员网大版题库.xlsx',True)
